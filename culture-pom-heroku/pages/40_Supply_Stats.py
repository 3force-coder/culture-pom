import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import date, datetime, timedelta
import calendar
import re

from auth import require_access, is_admin
from components import show_footer
from database import get_connection

# ============================================================
# CONFIGURATION PAGE
# ============================================================
st.set_page_config(
    page_title="Stats Supply — POMI",
    page_icon="🚛",
    layout="wide"
)

st.markdown("""<style>
    .block-container {padding-top:2rem!important;padding-bottom:0.5rem!important;
        padding-left:2rem!important;padding-right:2rem!important;}
    h1,h2,h3,h4{margin-top:0.3rem!important;margin-bottom:0.3rem!important;}
    [data-testid="stMetricValue"]{font-size:1.4rem!important;}
    hr{margin-top:0.5rem!important;margin-bottom:0.5rem!important;}
    .kpi-vert  { background:#f0faf0; border-left:4px solid #AFCA0A; border-radius:6px; padding:12px 16px; margin-bottom:8px; }
    .badge-previsionnel { background:#fff3cd; color:#856404; padding:2px 8px; border-radius:12px; font-size:0.8em; font-weight:600; }
    .badge-conditionnel { background:#f8d7da; color:#721c24; padding:2px 8px; border-radius:12px; font-size:0.8em; font-weight:600; }
    .info-box  { background:#e8f4f8; border:1px solid #bee5eb; border-radius:6px; padding:10px 14px; margin:8px 0; font-size:0.9em; }
</style>""", unsafe_allow_html=True)

require_access("COMMERCIAL")
st.title("🚛 Supply — Planning Appro & Stats Transports")

# ============================================================
# FILTRES GLOBAUX + COMPARAISON (identique pattern Frulog)
# ============================================================

def campagne_dates(year):
    return date(year - 1, 6, 1), date(year, 5, 31)

def get_campagnes():
    now = datetime.now()
    cy = now.year + 1 if now.month >= 6 else now.year
    return list(range(cy + 1, 2022, -1))

campagnes = get_campagnes()

with st.container():
    fc1, fc2, fc3, fc4 = st.columns([2, 1.5, 1.5, 2])
    with fc1:
        camp_opts = [0] + campagnes
        camp_labels_map = {0: "🔓 Toutes", **{y: f"Camp. {y} ({y-1}/06→{y}/05)" for y in campagnes}}
        sel_camp = st.selectbox("🗓️ Campagne", camp_opts, format_func=lambda y: camp_labels_map[y], key="f_camp")
    with fc2:
        d_min_camp = campagne_dates(sel_camp)[0] if sel_camp else date(2022, 1, 1)
        d_max_camp = campagne_dates(sel_camp)[1] if sel_camp else date(2027, 12, 31)
        date_deb = st.date_input("📅 Du", value=d_min_camp, min_value=d_min_camp, max_value=d_max_camp, key="f_deb")
    with fc3:
        date_fin = st.date_input("📅 Au", value=d_max_camp, min_value=d_min_camp, max_value=d_max_camp, key="f_fin")
    with fc4:
        mode_compare = st.selectbox("🔀 Comparaison", [
            "Aucune", "📅 Campagne vs Campagne", "📆 Mois vs Mois", "📆 Semaine vs Semaine"
        ], key="f_compare")

COMP_DATES = None
if mode_compare == "📅 Campagne vs Campagne":
    cc1, cc2 = st.columns(2)
    with cc1: c1 = st.selectbox("Campagne A", campagnes, key="comp_c1")
    with cc2: c2 = st.selectbox("Campagne B", [c for c in campagnes if c != c1], key="comp_c2")
    d1a, d1b = campagne_dates(c1); d2a, d2b = campagne_dates(c2)
    COMP_DATES = (f"Camp. {c1}", d1a, d1b, f"Camp. {c2}", d2a, d2b)
elif mode_compare == "📆 Mois vs Mois":
    mois_names = {1:'Janvier',2:'Février',3:'Mars',4:'Avril',5:'Mai',6:'Juin',
                  7:'Juillet',8:'Août',9:'Septembre',10:'Octobre',11:'Novembre',12:'Décembre'}
    cm1, cm2, cm3, cm4 = st.columns(4)
    with cm1: ma = st.selectbox("Mois A", range(1,13), format_func=lambda m: mois_names[m], key="comp_ma")
    with cm2: ya = st.number_input("Année A", 2022, 2027, datetime.now().year, key="comp_ya")
    with cm3: mb = st.selectbox("Mois B", range(1,13), format_func=lambda m: mois_names[m], index=max(0, datetime.now().month-2), key="comp_mb")
    with cm4: yb = st.number_input("Année B", 2022, 2027, datetime.now().year - 1, key="comp_yb")
    d1a = date(int(ya), int(ma), 1); d1b = date(int(ya), int(ma), calendar.monthrange(int(ya), int(ma))[1])
    d2a = date(int(yb), int(mb), 1); d2b = date(int(yb), int(mb), calendar.monthrange(int(yb), int(mb))[1])
    COMP_DATES = (f"{mois_names[ma]} {ya}", d1a, d1b, f"{mois_names[mb]} {yb}", d2a, d2b)
elif mode_compare == "📆 Semaine vs Semaine":
    cs1, cs2, cs3, cs4 = st.columns(4)
    with cs1: sa = st.number_input("Semaine A", 1, 53, max(1, datetime.now().isocalendar()[1]-1), key="comp_sa")
    with cs2: ysa = st.number_input("Année A", 2022, 2027, datetime.now().year, key="comp_ysa")
    with cs3: sb = st.number_input("Semaine B", 1, 53, max(1, datetime.now().isocalendar()[1]-1), key="comp_sb")
    with cs4: ysb = st.number_input("Année B", 2022, 2027, datetime.now().year - 1, key="comp_ysb")
    d1a = date.fromisocalendar(int(ysa), int(sa), 1); d1b = d1a + timedelta(days=6)
    d2a = date.fromisocalendar(int(ysb), int(sb), 1); d2b = d2a + timedelta(days=6)
    COMP_DATES = (f"S{sa}/{ysa}", d1a, d1b, f"S{sb}/{ysb}", d2a, d2b)

st.markdown("---")
DATE_DEB = date_deb
DATE_FIN = date_fin

# ============================================================
# CONSTANTES NORMALISATION SITES
# ============================================================

# Sites canoniques et leurs variantes connues + fuzzy
SITES_CANONIQUES = {
    "Saint-Flavy": [
        "SAINT FLAVY", "ST FLAVY", "SAINT-FLAVY", "STFLAVY",
        "SAINT  FLAVY", "S FLAVY", "ST-FLAVY"
    ],
    "Corroy": [
        "CORROY", "COROI", "CORROI"
    ],
    "La Motte Tilly": [
        "LA MOTTE TILLY", "LA MOTTE-TILLY", "LAMOTTE TILLY",
        "LA MOTTE TILI", "MOTTE TILLY"
    ],
    "Rougeaux": [
        "ROUGEAUX", "ROUGEAU", "ROUGEAUX (DEPOT)"
    ],
    "Pom Val": [
        "POM VAL", "POMVAL"
    ],
}

def normaliser_site(valeur_brute: str) -> str:
    """
    Normalise un nom de site de livraison par fuzzy matching.
    Retourne le nom canonique ou 'Autre' si aucun match suffisant.
    """
    if not valeur_brute or pd.isna(valeur_brute):
        return "Non renseigné"

    val = str(valeur_brute).upper().strip()

    # 1. Correspondance exacte parmi les variantes connues
    for canon, variantes in SITES_CANONIQUES.items():
        if val in variantes or val == canon.upper():
            return canon

    # 2. Fuzzy matching simple : ratio de caractères communs
    def similarity(a: str, b: str) -> float:
        """Ratio de Jaccard sur bigrammes."""
        def bigrams(s):
            return set(s[i:i+2] for i in range(len(s)-1))
        bg_a = bigrams(a)
        bg_b = bigrams(b)
        if not bg_a or not bg_b:
            return 0.0
        inter = bg_a & bg_b
        union = bg_a | bg_b
        return len(inter) / len(union)

    best_canon = None
    best_score = 0.0
    for canon, variantes in SITES_CANONIQUES.items():
        # Comparer avec le nom canonique et ses variantes
        candidats = [canon.upper()] + variantes
        for c in candidats:
            score = similarity(val, c)
            if score > best_score:
                best_score = score
                best_canon = canon

    if best_score >= 0.60:
        return best_canon

    return "Autre"


# ============================================================
# LOGIQUE DATE / STATUT
# ============================================================

def parser_date_statut(val_date, today: date):
    """
    Retourne (date_transport, semaine_text, statut, annee_semaine).
    """
    # Cas 1 : vraie date
    if isinstance(val_date, (datetime, pd.Timestamp)):
        d = val_date.date() if hasattr(val_date, 'date') else val_date
        if d <= today:
            statut = "realise"
        else:
            statut = "previsionnel"
        iso = d.isocalendar()
        annee_sem = f"{iso[0]}-S{iso[1]:02d}"
        return d, None, statut, annee_sem

    if isinstance(val_date, date):
        d = val_date
        statut = "realise" if d <= today else "previsionnel"
        iso = d.isocalendar()
        annee_sem = f"{iso[0]}-S{iso[1]:02d}"
        return d, None, statut, annee_sem

    # Cas 2 : texte type "S11", "S12", "S 11"
    if isinstance(val_date, str):
        match = re.match(r'^S\s*(\d{1,2})$', val_date.strip().upper())
        if match:
            num_sem = int(match.group(1))
            annee = today.year
            # Si num_sem < semaine courante → probablement année suivante
            # (cas rare en pratique — on laisse l'année courante)
            sem_text = val_date.strip().upper()
            annee_sem = f"{annee}-S{num_sem:02d}"
            return None, sem_text, "previsionnel", annee_sem

        # Cas 3 : autre texte
        return None, str(val_date), "conditionnel", None

    return None, None, "conditionnel", None


# ============================================================
# IMPORT EXCEL
# ============================================================

def lire_onglet(df_raw: pd.DataFrame, today: date) -> pd.DataFrame:
    """
    Nettoie et transforme un onglet du fichier Excel en DataFrame normalisé.
    """
    # Colonnes attendues
    col_map = {
        'JOUR': 'jour',
        'DATE': 'date_brute',
        'TRANSP': 'transporteur',
        'CHAUFFEUR': 'chauffeur',
        'SITE CHARGEMENT': 'site_chargement',
        'QUOI ?': 'quoi',
        'SITE LIVRAISON': 'site_livraison_brut',
        'CONDI': 'condi',
        'INFOS': 'infos',
        "HEURE D'ARRIVEE SUR SITE": 'heure_arrivee',
        'DESTINATION': 'destination',
        'AGREAGE': 'agreage',
    }

    # Renommer les colonnes présentes
    df = df_raw.copy()
    present = {k: v for k, v in col_map.items() if k in df.columns}
    df = df.rename(columns=present)

    # Supprimer lignes entièrement vides
    df = df.dropna(how='all')

    # Garder uniquement les lignes ayant au moins transporteur ou chauffeur
    key_cols = [c for c in ['transporteur', 'chauffeur', 'site_chargement'] if c in df.columns]
    if key_cols:
        df = df.dropna(subset=key_cols, how='all')

    if df.empty:
        return pd.DataFrame()

    rows = []
    for _, row in df.iterrows():
        val_date = row.get('date_brute', None)
        date_t, sem_text, statut, annee_sem = parser_date_statut(val_date, today)

        site_brut = str(row.get('site_livraison_brut', '') or '').strip()
        site_norm = normaliser_site(site_brut)

        rows.append({
            'jour':                 str(row.get('jour', '') or '').strip() or None,
            'date_transport':       date_t,
            'semaine_text':         sem_text,
            'statut':               statut,
            'transporteur':         str(row.get('transporteur', '') or '').strip() or None,
            'chauffeur':            str(row.get('chauffeur', '') or '').strip() or None,
            'site_chargement':      str(row.get('site_chargement', '') or '').strip() or None,
            'quoi':                 str(row.get('quoi', '') or '').strip() or None,
            'site_livraison_brut':  site_brut or None,
            'site_livraison_norm':  site_norm,
            'condi':                str(row.get('condi', '') or '').strip() or None,
            'infos':                str(row.get('infos', '') or '').strip() or None,
            'heure_arrivee':        str(row.get('heure_arrivee', '') or '').strip() or None,
            'destination':          str(row.get('destination', '') or '').strip() or None,
            'agreage':              str(row.get('agreage', '') or '').strip() or None,
            'annee_semaine':        annee_sem,
        })

    return pd.DataFrame(rows)


def _lire_onglet_openpyxl(ws, col_map: dict) -> pd.DataFrame:
    """
    Lit un worksheet openpyxl en mode read_only et retourne un DataFrame
    avec uniquement les colonnes définies dans col_map.
    Évite de charger le formatage et les cellules vides (économie mémoire).
    """
    rows_iter = ws.iter_rows(values_only=True)

    # Première ligne = en-têtes
    try:
        headers = [str(h).strip() if h is not None else '' for h in next(rows_iter)]
    except StopIteration:
        return pd.DataFrame()

    # Index des colonnes utiles uniquement
    col_indices = {}
    for src_name, dst_name in col_map.items():
        if src_name in headers:
            col_indices[headers.index(src_name)] = dst_name

    if not col_indices:
        return pd.DataFrame()

    records = []
    for row in rows_iter:
        # Ignorer les lignes entièrement vides
        if all(v is None for v in row):
            continue
        record = {dst: row[idx] if idx < len(row) else None
                  for idx, dst in col_indices.items()}
        records.append(record)

    return pd.DataFrame(records)


def _upsert_supply_df(df_all: pd.DataFrame, username: str):
    """
    Upsert en BDD un DataFrame déjà normalisé (issu du cache session_state).
    Retourne (ok, message, nb_inserts, nb_skips).
    """
    conn = get_connection()
    if not conn:
        return False, "Connexion BDD impossible.", 0, 0

    nb_insert = 0
    nb_skip = 0

    try:
        cursor = conn.cursor()

        for _, r in df_all.iterrows():
            try:
                cursor.execute("""
                    INSERT INTO supply_transports
                        (jour, date_transport, semaine_text, statut,
                         transporteur, chauffeur, site_chargement, quoi,
                         site_livraison_brut, site_livraison_norm,
                         condi, infos, heure_arrivee, destination, agreage,
                         annee_semaine, imported_by, imported_at)
                    VALUES
                        (%s, %s, %s, %s,
                         %s, %s, %s, %s,
                         %s, %s,
                         %s, %s, %s, %s, %s,
                         %s, %s, CURRENT_TIMESTAMP)
                    ON CONFLICT (date_transport, transporteur, chauffeur, site_chargement, quoi, site_livraison_brut)
                    DO UPDATE SET
                        statut               = EXCLUDED.statut,
                        site_livraison_norm  = EXCLUDED.site_livraison_norm,
                        semaine_text         = EXCLUDED.semaine_text,
                        condi                = EXCLUDED.condi,
                        infos                = EXCLUDED.infos,
                        destination          = EXCLUDED.destination,
                        annee_semaine        = EXCLUDED.annee_semaine,
                        imported_by          = EXCLUDED.imported_by,
                        imported_at          = CURRENT_TIMESTAMP
                    RETURNING (xmax = 0) AS inserted
                """, (
                    r['jour'],
                    r['date_transport'],
                    r['semaine_text'],
                    r['statut'],
                    r['transporteur'],
                    r['chauffeur'],
                    r['site_chargement'],
                    r['quoi'],
                    r['site_livraison_brut'],
                    r['site_livraison_norm'],
                    r['condi'],
                    r['infos'],
                    r['heure_arrivee'],
                    r['destination'],
                    r['agreage'],
                    r['annee_semaine'],
                    username,
                ))
                result = cursor.fetchone()
                if result and result['inserted']:
                    nb_insert += 1
                else:
                    nb_skip += 1
            except Exception:
                nb_skip += 1
                conn.rollback()

        conn.commit()
        cursor.close()
        conn.close()

        msg = (f"✅ Import OK — {nb_insert} nouvelles lignes / {nb_skip} mises à jour ou ignorées")
        return True, msg, nb_insert, nb_skip

    except Exception as e:
        conn.rollback()
        conn.close()
        return False, f"Erreur BDD : {str(e)}", 0, 0


def importer_supply(uploaded_file, username: str):
    """
    Lit les onglets 'Planning appro' et 'Archives' du fichier Excel,
    fusionne, normalise et upsert en BDD.
    Retourne (ok, message, nb_inserts, nb_skips).
    Utilise openpyxl read_only=True pour éviter l'explosion mémoire
    sur les fichiers Archives volumineux.
    """
    import openpyxl
    import io

    today = date.today()
    dfs = []
    onglets_lus = []

    # Mapping colonnes Excel → noms internes
    col_map = {
        'JOUR':                    'jour',
        'DATE':                    'date_brute',
        'TRANSP':                  'transporteur',
        'CHAUFFEUR':               'chauffeur',
        'SITE CHARGEMENT':         'site_chargement',
        'QUOI ?':                  'quoi',
        'SITE LIVRAISON':          'site_livraison_brut',
        'CONDI':                   'condi',
        'INFOS':                   'infos',
        "HEURE D'ARRIVEE SUR SITE": 'heure_arrivee',
        'DESTINATION':             'destination',
        'AGREAGE':                 'agreage',
    }

    try:
        # Lire le fichier en mémoire une seule fois
        file_bytes = uploaded_file.read()
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            read_only=True,   # pas de chargement du formatage
            data_only=True,   # valeurs calculées, pas les formules
        )

        for onglet in ['Planning appro', 'Archives']:
            if onglet not in wb.sheetnames:
                continue
            ws = wb[onglet]
            df_raw = _lire_onglet_openpyxl(ws, col_map)
            df_norm = lire_onglet(df_raw, today)
            if not df_norm.empty:
                dfs.append(df_norm)
                onglets_lus.append(onglet)

        wb.close()

        if not dfs:
            return False, "Aucune donnée exploitable trouvée dans les onglets.", 0, 0

        df_all = pd.concat(dfs, ignore_index=True)
        df_all = df_all.dropna(subset=['transporteur', 'chauffeur'], how='all')

        if df_all.empty:
            return False, "Aucune ligne valide (transporteur ou chauffeur manquant).", 0, 0

    except Exception as e:
        return False, f"Erreur lecture Excel : {str(e)}", 0, 0

    # Upsert BDD
    conn = get_connection()
    if not conn:
        return False, "Connexion BDD impossible.", 0, 0

    nb_insert = 0
    nb_skip = 0

    try:
        cursor = conn.cursor()

        for _, r in df_all.iterrows():
            try:
                cursor.execute("""
                    INSERT INTO supply_transports
                        (jour, date_transport, semaine_text, statut,
                         transporteur, chauffeur, site_chargement, quoi,
                         site_livraison_brut, site_livraison_norm,
                         condi, infos, heure_arrivee, destination, agreage,
                         annee_semaine, imported_by, imported_at)
                    VALUES
                        (%s, %s, %s, %s,
                         %s, %s, %s, %s,
                         %s, %s,
                         %s, %s, %s, %s, %s,
                         %s, %s, CURRENT_TIMESTAMP)
                    ON CONFLICT (date_transport, transporteur, chauffeur, site_chargement, quoi, site_livraison_brut)
                    DO UPDATE SET
                        statut               = EXCLUDED.statut,
                        site_livraison_norm  = EXCLUDED.site_livraison_norm,
                        semaine_text         = EXCLUDED.semaine_text,
                        condi                = EXCLUDED.condi,
                        infos                = EXCLUDED.infos,
                        destination          = EXCLUDED.destination,
                        annee_semaine        = EXCLUDED.annee_semaine,
                        imported_by          = EXCLUDED.imported_by,
                        imported_at          = CURRENT_TIMESTAMP
                    RETURNING (xmax = 0) AS inserted
                """, (
                    r['jour'],
                    r['date_transport'],
                    r['semaine_text'],
                    r['statut'],
                    r['transporteur'],
                    r['chauffeur'],
                    r['site_chargement'],
                    r['quoi'],
                    r['site_livraison_brut'],
                    r['site_livraison_norm'],
                    r['condi'],
                    r['infos'],
                    r['heure_arrivee'],
                    r['destination'],
                    r['agreage'],
                    r['annee_semaine'],
                    username,
                ))
                result = cursor.fetchone()
                if result and result['inserted']:
                    nb_insert += 1
                else:
                    nb_skip += 1
            except Exception:
                nb_skip += 1
                conn.rollback()

        conn.commit()
        cursor.close()
        conn.close()

        msg = (f"✅ Import OK — onglets : {', '.join(onglets_lus)}\n"
               f"{nb_insert} nouvelles lignes / {nb_skip} mises à jour ou ignorées")
        return True, msg, nb_insert, nb_skip

    except Exception as e:
        conn.rollback()
        conn.close()
        return False, f"Erreur BDD : {str(e)}", 0, 0


# ============================================================
# REQUÊTES BDD
# ============================================================

@st.cache_data(ttl=120)
def get_realise(date_debut=None, date_fin=None):
    conn = get_connection()
    if not conn:
        return pd.DataFrame()
    try:
        cursor = conn.cursor()
        params = []
        where = "WHERE statut = 'realise'"
        if date_debut:
            where += " AND date_transport >= %s"
            params.append(date_debut)
        if date_fin:
            where += " AND date_transport <= %s"
            params.append(date_fin)

        cursor.execute(f"""
            SELECT
                id, jour, date_transport, transporteur, chauffeur,
                site_chargement, quoi, site_livraison_brut, site_livraison_norm,
                condi, infos, annee_semaine, imported_at
            FROM supply_transports
            {where}
            ORDER BY date_transport DESC
        """, params)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        if not rows:
            return pd.DataFrame()
        return pd.DataFrame([dict(r) for r in rows])
    except Exception:
        conn.close()
        return pd.DataFrame()


@st.cache_data(ttl=120)
def get_previsionnel():
    conn = get_connection()
    if not conn:
        return pd.DataFrame()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                id, jour, date_transport, semaine_text, statut,
                transporteur, chauffeur, site_chargement, quoi,
                site_livraison_brut, site_livraison_norm,
                condi, infos, annee_semaine, destination
            FROM supply_transports
            WHERE statut IN ('previsionnel', 'conditionnel')
            ORDER BY annee_semaine, date_transport, chauffeur
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        if not rows:
            return pd.DataFrame()
        return pd.DataFrame([dict(r) for r in rows])
    except Exception:
        conn.close()
        return pd.DataFrame()


@st.cache_data(ttl=120)
def get_semaines_dispo():
    conn = get_connection()
    if not conn:
        return []
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT annee_semaine
            FROM supply_transports
            WHERE statut = 'realise' AND annee_semaine IS NOT NULL
            ORDER BY annee_semaine DESC
            LIMIT 20
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return [r['annee_semaine'] for r in rows]
    except Exception:
        conn.close()
        return []


# ============================================================
# COMPOSANTS D'ANALYSE
# ============================================================

def kpis_row(df: pd.DataFrame, label_periode: str = ""):
    """Affiche une ligne de KPIs pour un DataFrame de transports."""
    nb_total = len(df)
    nb_vrac = len(df[df['condi'].str.upper() == 'VRAC']) if 'condi' in df.columns else 0
    nb_palox = len(df[df['condi'].str.upper() == 'PALOX']) if 'condi' in df.columns else 0
    nb_chauf = df['chauffeur'].nunique() if 'chauffeur' in df.columns else 0

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("🚛 Voyages total", nb_total, help=label_periode)
    with c2:
        st.metric("🌾 VRAC", nb_vrac)
    with c3:
        st.metric("📦 PALOX", nb_palox)
    with c4:
        st.metric("👤 Chauffeurs", nb_chauf)


def graphe_hebdo(df: pd.DataFrame, titre: str):
    """Évolution hebdomadaire du nombre de voyages."""
    if df.empty or 'annee_semaine' not in df.columns:
        st.info("Pas de données pour le graphique hebdomadaire.")
        return

    grp = df.groupby('annee_semaine').size().reset_index(name='nb_voyages')
    grp = grp.sort_values('annee_semaine')

    fig = px.bar(
        grp, x='annee_semaine', y='nb_voyages',
        title=titre,
        labels={'annee_semaine': 'Semaine', 'nb_voyages': 'Nb voyages'},
        color_discrete_sequence=['#AFCA0A'],
        text='nb_voyages',
    )
    fig.update_traces(textposition='outside')
    fig.update_layout(
        plot_bgcolor='white', paper_bgcolor='white',
        xaxis_tickangle=-45, height=350, margin=dict(t=40, b=60)
    )
    st.plotly_chart(fig, use_container_width=True)


def graphe_chauffeur(df: pd.DataFrame):
    """Répartition par chauffeur."""
    if df.empty or 'chauffeur' not in df.columns:
        return

    grp = df.groupby('chauffeur').size().reset_index(name='nb_voyages')
    grp = grp.sort_values('nb_voyages', ascending=True)

    fig = px.bar(
        grp, x='nb_voyages', y='chauffeur', orientation='h',
        title="Voyages par chauffeur",
        labels={'nb_voyages': 'Nb voyages', 'chauffeur': 'Chauffeur'},
        color_discrete_sequence=['#AFCA0A'],
        text='nb_voyages',
    )
    fig.update_traces(textposition='outside')
    fig.update_layout(
        plot_bgcolor='white', paper_bgcolor='white',
        height=max(250, len(grp) * 40 + 80),
        margin=dict(t=40, b=20)
    )
    st.plotly_chart(fig, use_container_width=True)


def graphe_site_arrivee(df: pd.DataFrame):
    """Répartition par site de livraison normalisé."""
    if df.empty or 'site_livraison_norm' not in df.columns:
        return

    grp = df.groupby('site_livraison_norm').size().reset_index(name='nb_voyages')
    grp = grp.sort_values('nb_voyages', ascending=False)

    fig = px.pie(
        grp, values='nb_voyages', names='site_livraison_norm',
        title="Répartition par site d'arrivée",
        color_discrete_sequence=px.colors.qualitative.Set2,
        hole=0.35,
    )
    fig.update_traces(textinfo='percent+label+value')
    fig.update_layout(height=350, margin=dict(t=40, b=20))
    st.plotly_chart(fig, use_container_width=True)


def tableau_sites_arrivee(df: pd.DataFrame):
    """Tableau détaillé par site normalisé."""
    if df.empty:
        return

    grp = df.groupby(['site_livraison_norm', 'condi']).size().reset_index(name='nb_voyages')
    pivot = grp.pivot_table(
        index='site_livraison_norm', columns='condi',
        values='nb_voyages', fill_value=0, aggfunc='sum'
    ).reset_index()
    pivot.columns.name = None
    pivot['Total'] = pivot.select_dtypes('number').sum(axis=1)
    pivot = pivot.sort_values('Total', ascending=False)
    pivot = pivot.rename(columns={'site_livraison_norm': 'Site d\'arrivée'})

    st.dataframe(pivot, use_container_width=True, hide_index=True)


def comparaison_semaines(df: pd.DataFrame, sem_ref: str, sem_comp: str):
    """Compare deux semaines côte-à-côte."""
    d_ref = df[df['annee_semaine'] == sem_ref]
    d_comp = df[df['annee_semaine'] == sem_comp] if sem_comp else pd.DataFrame()

    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f"**{sem_ref}** — {len(d_ref)} voyages")
        if not d_ref.empty:
            st.dataframe(
                d_ref[['chauffeur', 'site_chargement', 'quoi', 'site_livraison_norm', 'condi']],
                use_container_width=True, hide_index=True
            )

    with col2:
        if not d_comp.empty:
            st.markdown(f"**{sem_comp}** — {len(d_comp)} voyages")
            st.dataframe(
                d_comp[['chauffeur', 'site_chargement', 'quoi', 'site_livraison_norm', 'condi']],
                use_container_width=True, hide_index=True
            )
        elif sem_comp:
            st.markdown(f"**{sem_comp}** — pas de données")
            st.info("Aucune donnée pour cette semaine.")


# ============================================================
# ONGLETS PRINCIPAUX
# ============================================================

tab_import, tab_realise, tab_previ = st.tabs([
    "📥 Import",
    "📅 Réalisé",
    "🔮 Prévisionnel",
])

# ────────────────────────────────────────────────────────────
# ONGLET 1 : IMPORT
# ────────────────────────────────────────────────────────────
with tab_import:
    st.subheader("📥 Import Planning Appro")

    st.markdown("""
    <div class="info-box">
    📌 Importez le fichier <strong>PLANNING_APPRO.xlsx</strong>.<br>
    Les deux onglets <em>Planning appro</em> (semaine courante + prévi) et <em>Archives</em> (historique)
    sont lus automatiquement en un seul import.<br>
    Les doublons sont ignorés, les lignes existantes mises à jour.
    </div>
    """, unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "Sélectionner le fichier Excel",
        type=['xlsx', 'xls'],
        key="supply_upload"
    )

    if uploaded:
        # ── Lecture unique avec openpyxl read_only ──────────────
        # On lit le fichier une seule fois et on stocke le DataFrame
        # en session_state pour éviter une double lecture mémoire.
        import openpyxl, io

        file_id = f"{uploaded.name}_{uploaded.size}"
        if st.session_state.get('supply_file_id') != file_id:
            # Nouveau fichier : lire et mettre en cache
            with st.spinner("Lecture du fichier en cours…"):
                try:
                    file_bytes = uploaded.read()
                    wb = openpyxl.load_workbook(
                        io.BytesIO(file_bytes),
                        read_only=True,
                        data_only=True,
                    )
                    col_map = {
                        'JOUR': 'jour', 'DATE': 'date_brute',
                        'TRANSP': 'transporteur', 'CHAUFFEUR': 'chauffeur',
                        'SITE CHARGEMENT': 'site_chargement', 'QUOI ?': 'quoi',
                        'SITE LIVRAISON': 'site_livraison_brut', 'CONDI': 'condi',
                        'INFOS': 'infos', "HEURE D'ARRIVEE SUR SITE": 'heure_arrivee',
                        'DESTINATION': 'destination', 'AGREAGE': 'agreage',
                    }
                    today = date.today()
                    dfs_prev = []
                    onglets_lus = []
                    for ong in ['Planning appro', 'Archives']:
                        if ong in wb.sheetnames:
                            df_raw = _lire_onglet_openpyxl(wb[ong], col_map)
                            df_norm = lire_onglet(df_raw, today)
                            if not df_norm.empty:
                                dfs_prev.append(df_norm)
                                onglets_lus.append(ong)
                    wb.close()

                    if dfs_prev:
                        df_cache = pd.concat(dfs_prev, ignore_index=True).dropna(
                            subset=['transporteur', 'chauffeur'], how='all'
                        )
                        st.session_state['supply_df_cache']   = df_cache
                        st.session_state['supply_onglets']    = onglets_lus
                        st.session_state['supply_file_id']    = file_id
                    else:
                        st.error("Aucune donnée exploitable dans les onglets.")
                        st.stop()

                except Exception as e:
                    st.error(f"Erreur lecture fichier : {str(e)}")
                    st.stop()

        # ── Affichage aperçu depuis cache ────────────────────────
        df_apercu  = st.session_state.get('supply_df_cache', pd.DataFrame())
        ong_lus    = st.session_state.get('supply_onglets', [])

        if not df_apercu.empty:
            st.caption(f"Onglets lus : {', '.join(ong_lus)}")

            nb_r = len(df_apercu[df_apercu['statut'] == 'realise'])
            nb_p = len(df_apercu[df_apercu['statut'] == 'previsionnel'])
            nb_c = len(df_apercu[df_apercu['statut'] == 'conditionnel'])

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Lignes total",      len(df_apercu))
            c2.metric("✅ Réalisé",         nb_r)
            c3.metric("🔮 Prévisionnel",    nb_p)
            c4.metric("⚠️ Conditionnel",    nb_c)

            with st.expander("Aperçu des données (20 premières lignes)"):
                cols_show = ['statut', 'annee_semaine', 'date_transport', 'semaine_text',
                             'transporteur', 'chauffeur', 'site_chargement', 'quoi',
                             'site_livraison_norm', 'condi']
                cols_show = [c for c in cols_show if c in df_apercu.columns]
                st.dataframe(df_apercu[cols_show].head(20),
                             use_container_width=True, hide_index=True)

            if st.button("🚀 Importer dans la base", type="primary", use_container_width=True):
                with st.spinner("Import en cours…"):
                    # Upsert direct depuis le DataFrame en cache (pas de relecture fichier)
                    ok, msg, nb_ins, nb_sk = _upsert_supply_df(
                        df_apercu,
                        st.session_state.get('username', '?')
                    )
                if ok:
                    st.success(msg)
                    st.cache_data.clear()
                    # Vider le cache fichier pour forcer relecture au prochain upload
                    st.session_state.pop('supply_file_id', None)
                else:
                    st.error(msg)

    # Dernier import
    try:
        conn_info = get_connection()
        if conn_info:
            cur = conn_info.cursor()
            cur.execute("""
                SELECT imported_by, imported_at, COUNT(*) as nb
                FROM supply_transports
                GROUP BY imported_by, imported_at
                ORDER BY imported_at DESC
                LIMIT 1
            """)
            last = cur.fetchone()
            cur.close()
            conn_info.close()
            if last:
                st.caption(
                    f"Dernier import : {last['imported_at'].strftime('%d/%m/%Y %H:%M')} "
                    f"par {last['imported_by']} — {last['nb']} lignes"
                )
    except Exception:
        pass


# ────────────────────────────────────────────────────────────
# ONGLET 2 : RÉALISÉ
# ────────────────────────────────────────────────────────────
with tab_realise:
    st.subheader("📅 Transports réalisés")

    if COMP_DATES:
        la, d1a, d1b, lb, d2a, d2b = COMP_DATES
        st.info(f"🔀 Comparaison : **{la}** vs **{lb}**")

        filtre_condi = st.selectbox("Type de transport", ["Tous", "VRAC", "PALOX"], key="r_condi_comp")

        df_A = get_realise(d1a, d1b)
        df_B = get_realise(d2a, d2b)
        if filtre_condi != "Tous":
            df_A = df_A[df_A['condi'].str.upper() == filtre_condi] if not df_A.empty else df_A
            df_B = df_B[df_B['condi'].str.upper() == filtre_condi] if not df_B.empty else df_B

        col_A, col_B = st.columns(2)
        with col_A:
            st.markdown(f"### {la}")
            kpis_row(df_A, la)
            graphe_hebdo(df_A, f"Voyages / semaine — {la}")
            graphe_chauffeur(df_A)
            graphe_site_arrivee(df_A)
        with col_B:
            st.markdown(f"### {lb}")
            kpis_row(df_B, lb)
            graphe_hebdo(df_B, f"Voyages / semaine — {lb}")
            graphe_chauffeur(df_B)
            graphe_site_arrivee(df_B)

    else:
        # Mode standard : filtre date libre
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            date_debut = st.date_input("Depuis le", value=DATE_DEB, key="r_debut")
        with col_f2:
            date_fin_r = st.date_input("Jusqu'au", value=DATE_FIN, key="r_fin")
        with col_f3:
            filtre_condi = st.selectbox("Type de transport", ["Tous", "VRAC", "PALOX"], key="r_condi")

        df_real = get_realise(date_debut, date_fin_r)

    if df_real.empty:
        st.info("Aucune donnée réalisée pour cette période. Effectuez un import.")
    else:
        # Filtre condi
        if filtre_condi != "Tous":
            df_real = df_real[
                df_real['condi'].str.upper() == filtre_condi
            ]

        st.markdown("---")
        kpis_row(df_real, f"{date_debut} → {date_fin}")
        st.markdown("---")

        col_g1, col_g2 = st.columns([3, 2])
        with col_g1:
            graphe_hebdo(df_real, "📈 Évolution hebdomadaire des voyages")
        with col_g2:
            graphe_chauffeur(df_real)

        st.markdown("---")
        st.subheader("🏭 Analyse par site d'arrivée")

        col_s1, col_s2 = st.columns([2, 3])
        with col_s1:
            graphe_site_arrivee(df_real)
        with col_s2:
            tableau_sites_arrivee(df_real)

        st.markdown("---")
        st.subheader("🔀 Comparaison semaine à semaine")

        semaines = get_semaines_dispo()
        if len(semaines) >= 2:
            col_c1, col_c2 = st.columns(2)
            with col_c1:
                sem_ref = st.selectbox("Semaine de référence", semaines, index=0, key="sem_ref")
            with col_c2:
                sem_comp = st.selectbox(
                    "Semaine de comparaison",
                    ["(aucune)"] + semaines,
                    index=1 if len(semaines) > 1 else 0,
                    key="sem_comp"
                )
            sem_comp_val = None if sem_comp == "(aucune)" else sem_comp
            df_comp = get_realise()  # sans filtre date pour avoir tout l'historique
            comparaison_semaines(df_comp, sem_ref, sem_comp_val)
        else:
            st.info("Importez plusieurs semaines pour activer la comparaison.")

        # Tableau détail
        with st.expander("📋 Détail des lignes"):
            cols_detail = ['date_transport', 'annee_semaine', 'transporteur', 'chauffeur',
                           'site_chargement', 'quoi', 'site_livraison_norm',
                           'site_livraison_brut', 'condi', 'infos']
            cols_detail = [c for c in cols_detail if c in df_real.columns]
            st.dataframe(
                df_real[cols_detail].sort_values('date_transport', ascending=False),
                use_container_width=True, hide_index=True
            )


# ────────────────────────────────────────────────────────────
# ONGLET 3 : PRÉVISIONNEL
# ────────────────────────────────────────────────────────────
with tab_previ:
    st.subheader("🔮 Transports prévisionnels")

    st.markdown("""
    <div class="info-box">
    ⚠️ Les lignes <strong>prévisionnelles</strong> correspondent aux dates futures ou aux codes semaine
    (ex : <em>S11</em>) — elles ne sont pas encore confirmées.<br>
    Les lignes <strong>conditionnelles</strong> ont une indication libre (ex : <em>suivant essai</em>).
    </div>
    """, unsafe_allow_html=True)

    df_previ = get_previsionnel()

    if df_previ.empty:
        st.info("Aucun transport prévisionnel en base. Effectuez un import.")
    else:
        # KPIs
        nb_previ = len(df_previ[df_previ['statut'] == 'previsionnel'])
        nb_cond = len(df_previ[df_previ['statut'] == 'conditionnel'])
        semaines_previ = df_previ[df_previ['statut'] == 'previsionnel']['annee_semaine'].dropna().unique()

        c1, c2, c3 = st.columns(3)
        c1.metric("🔮 Prévisionnels", nb_previ)
        c2.metric("⚠️ Conditionnels", nb_cond)
        c3.metric("📅 Semaines concernées", len(semaines_previ))

        st.markdown("---")

        # Filtre par semaine
        toutes_sems = sorted(df_previ['annee_semaine'].dropna().unique())
        if toutes_sems:
            sem_filtre = st.multiselect(
                "Filtrer par semaine",
                options=toutes_sems,
                default=toutes_sems,
                key="previ_sem_filtre"
            )
            df_previ_f = df_previ[
                df_previ['annee_semaine'].isin(sem_filtre) |
                df_previ['annee_semaine'].isna()
            ]
        else:
            df_previ_f = df_previ

        # Tableau principal avec badge statut
        def fmt_statut(s):
            if s == 'previsionnel':
                return '🔮 Prévi'
            elif s == 'conditionnel':
                return '⚠️ Conditionnel'
            return s

        df_display = df_previ_f.copy()
        df_display['Statut'] = df_display['statut'].apply(fmt_statut)
        df_display['Semaine'] = df_display.apply(
            lambda r: r['annee_semaine'] if r['annee_semaine'] else r['semaine_text'], axis=1
        )
        df_display['Date'] = df_display.apply(
            lambda r: str(r['date_transport']) if r['date_transport'] else r['semaine_text'] or '', axis=1
        )

        cols_previ = ['Statut', 'Semaine', 'Date', 'transporteur', 'chauffeur',
                      'site_chargement', 'quoi', 'site_livraison_norm', 'condi', 'infos']
        cols_previ = [c for c in cols_previ if c in df_display.columns]

        st.dataframe(
            df_display[cols_previ].rename(columns={
                'transporteur': 'Transporteur',
                'chauffeur': 'Chauffeur',
                'site_chargement': 'Site chargement',
                'quoi': 'Quoi',
                'site_livraison_norm': 'Site arrivée',
                'condi': 'Condi',
                'infos': 'Infos',
            }),
            use_container_width=True,
            hide_index=True,
        )

        # Graphe répartition site arrivée pour les prévis
        st.markdown("---")
        col_p1, col_p2 = st.columns(2)
        with col_p1:
            graphe_site_arrivee(df_previ_f)
        with col_p2:
            graphe_chauffeur(df_previ_f)

# ============================================================
# FOOTER
# ============================================================
show_footer()
